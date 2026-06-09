from typing import Sequence, Union, List, Dict, Any, Tuple
from collections import Counter
from datetime import datetime, timedelta, timezone
from itertools import zip_longest
from PIL import Image
import calendar
import os
import glob
import random
import ast
import json
import base64
import pytz
import pyotp
from robot.libraries.BuiltIn import BuiltIn
from robot.utils import DotDict
from openpyxl import load_workbook
import zipfile
from robot.api import logger
import uuid
import yaml
import requests
import xlwings as xw
import win32com.client



class MyLibrary:
    """A Robot Framework library providing utilities for datetime, number, list, string, JSON, file, browser, dictionary, and security operations.

    This library encapsulates a collection of keywords designed for automation tasks, including datetime manipulation, list processing, JSON handling, file operations, browser interactions, and security features like OTP generation. It is intended to be used within Robot Framework test suites to enhance test automation capabilities.

    Examples:
        | Library | MyLibrary |
        | ${is_weekend} = | Is Weekend | 2025-06-01 | %Y-%m-%d |
        | ${json}= | Format JSON With Values | ${keys} | {"id": "837"} | id |
    """

    def is_weekend(self, date_str: str, timestamp_format: str) -> bool:
        """Check if a given date is a weekend (Sunday).

        Args:
            date_str: Date string to check (e.g., '2025-06-01').
            timestamp_format: Format of the date string (e.g., '%Y-%m-%d').

        Returns:
            bool: True if the date is Sunday, False otherwise.

        Raises:
            ValueError: If date_str cannot be parsed with the given format.

        Examples:
            | ${is_weekend}= | Is Weekend | 2025-06-01 | %Y-%m-%d |
        """
        try:
            date_obj = datetime.strptime(date_str, timestamp_format)
            return date_obj.weekday() == 6
        except ValueError as e:
            raise ValueError(f"Invalid date format for '{date_str}': {e}")

    def is_not_weekend(self, date_str: str, timestamp_format: str) -> bool:
        """Check if a given date is not a weekend (Sunday).

        Args:
            date_str: Date string to check (e.g., '2025-06-02').
            timestamp_format: Format of the date string (e.g., '%Y-%m-%d').

        Returns:
            bool: True if the date is not Sunday, False otherwise.

        Raises:
            ValueError: If date_str cannot be parsed with the given format.

        Examples:
            | ${is_not_weekend}= | Is Not Weekend | 2025-06-02 | %Y-%m-%d |
        """
        return not self.is_weekend(date_str, timestamp_format)

    def get_current_timestamp(self) -> int:
        """Get the current timestamp in seconds since epoch (UTC).

        Returns:
            int: Current timestamp in seconds.

        Examples:
            | ${timestamp}= | Get Current Timestamp |
        """
        return calendar.timegm(datetime.utcnow().utctimetuple())

    def convert_timestamp_to_datetime(self, timestamp: Union[str, int], format_str: str = '%Y-%m-%d %H:%M:%S') -> str:
        """Convert a timestamp (seconds or milliseconds) to a formatted datetime string.

        Args:
            timestamp: Timestamp value (e.g., 1734687527000 for milliseconds).
            format_str: Desired output format (default: '%Y-%m-%d %H:%M:%S').

        Returns:
            str: Formatted datetime string.

        Raises:
            ValueError: If timestamp is invalid.

        Examples:
            | ${date}= | Convert Timestamp To Datetime | 1734687527000 | %d/%m/%Y %H:%M |
            # Returns: '20/12/2024 16:38'
        """
        try:
            timestamp_float = float(timestamp) / 1000 if len(str(int(timestamp))) > 10 else float(timestamp)
            return datetime.fromtimestamp(timestamp_float).strftime(format_str)
        except (ValueError, TypeError) as e:
            raise ValueError(f"Invalid timestamp '{timestamp}': {e}")

    def convert_gmt_datetime(
        self,
        date_str: str,
        input_format: str = '%Y-%m-%dT%H:%M:%SZ',
        output_format: str = '%Y-%m-%d %H:%M:%S'
    ) -> str:
        """Convert a GMT datetime to GMT+7 timezone and format.

        Args:
            date_str: Input datetime string (e.g., '2021-06-08T08:39:49Z').
            input_format: Input format (default: '%Y-%m-%dT%H:%M:%SZ').
            output_format: Output format (default: '%Y-%m-%d %H:%M:%S').

        Returns:
            str: Formatted datetime string in GMT+7.

        Raises:
            ValueError: If date_str cannot be parsed.

        Examples:
            | ${result}= | Convert GMT Datetime | 2021-06-08T08:39:49Z |
        """
        try:
            local_tz = pytz.timezone("Asia/Ho_Chi_Minh")
            naive_dt = datetime.strptime(date_str, input_format)
            local_dt = local_tz.localize(naive_dt, is_dst=None)
            utc_dt = local_dt.astimezone(pytz.utc)
            return utc_dt.strftime(output_format)
        except ValueError as e:
            raise ValueError(f"Invalid datetime format for '{date_str}': {e}")

    def convert_datetime_to_date(
        self,
        date_str: str,
        input_format: str = '%Y-%m-%dT%H:%M:%SZ',
        output_format: str = '%d/%m/%Y'
    ) -> str:
        """Convert a datetime to a date string in GMT+7 timezone.

        Args:
            date_str: Input datetime string (e.g., '2021-06-08T08:39:49Z').
            input_format: Input format (default: '%Y-%m-%dT%H:%M:%SZ').
            output_format: Output format (default: '%d/%m/%Y').

        Returns:
            str: Formatted date string.

        Raises:
            ValueError: If date_str cannot be parsed.

        Examples:
            | ${date}= | Convert Datetime To Date | 2021-06-08T08:39:49Z |
        """
        return self.convert_gmt_datetime(date_str, input_format, output_format)

    def is_datetime_greater_or_equal(self, date1: str, date2: str, timestamp_format: str) -> bool:
        """Check if date1 is greater than or equal to date2.

        Args:
            date1: First date string.
            date2: Second date string.
            timestamp_format: Format of the date strings (e.g., '%d/%m/%Y %H:%M:%S').

        Returns:
            bool: True if date1 >= date2, False otherwise.

        Raises:
            ValueError: If dates cannot be parsed.

        Examples:
            | ${result}= | Is Datetime Greater Or Equal | 22/09/2022 18:01:30 | 30/09/2022 18:01:30 | %d/%m/%Y %H:%M:%S |
        """
        try:
            date1_obj = datetime.strptime(date1, timestamp_format)
            date2_obj = datetime.strptime(date2, timestamp_format)
            return date1_obj >= date2_obj
        except ValueError as e:
            raise ValueError(f"Invalid date format: {e}")

    def is_datetime_less_or_equal(self, date1: str, date2: str, timestamp_format: str) -> bool:
        """Check if date1 is less than or equal to date2.

        Args:
            date1: First date string.
            date2: Second date string.
            timestamp_format: Format of the date strings.

        Returns:
            bool: True if date1 <= date2, False otherwise.

        Raises:
            ValueError: If dates cannot be parsed.

        Examples:
            | ${result}= | Is Datetime Less Or Equal | 22/09/2022 18:01:30 | 30/09/2022 18:01:30 | %d/%m/%Y %H:%M:%S |
        """
        return not self.is_datetime_greater_or_equal(date1, date2, timestamp_format)

    def get_last_date_of_month(self, format_str: str = '%Y-%m-%d') -> str:
        """Get the last date of the current month.

        Args:
            format_str: Output format (default: '%Y-%m-%d').

        Returns:
            str: Formatted last date of the month.

        Examples:
            | ${last_date}= | Get Last Date Of Month | %d/%m/%Y |
        """
        today = datetime.today()
        year, month = today.year, today.month
        last_day = calendar.monthlen(year, month)
        return datetime(year, month, last_day).strftime(format_str)

    def get_first_date_of_month(self, format_str: str = '%Y-%m-%d') -> str:
        """Get the first date of the current month.

        Args:
            format_str: Output format (default: '%Y-%m-%d').

        Returns:
            str: Formatted first date of the month.

        Examples:
            | ${first_date}= | Get First Date Of Month | %d/%m/%Y |
        """
        today = datetime.today()
        return datetime(today.year, today.month, 1).strftime(format_str)
    
    def get_current_iso_datetime(self):
        tz = timezone(timedelta(hours=7))
        now = datetime.now(tz)
        iso = now.isoformat(timespec='milliseconds')  # sẽ tự thêm +07:00
        return iso
    
    def get_month_date_range(self, year: int, month: int) -> Tuple[str, str]:
        """Get the start and end dates of a specified month.

        Args:
            year: Year of the month (e.g., 2025).
            month: Month number (1-12).

        Returns:
            Tuple[str, str]: Start and end dates in 'YYYY-MM-DD HH:MM:SS.ssss' format.

        Raises:
            ValueError: If year or month is invalid.

        Examples:
            | ${start} | ${end}= | Get Month Date Range | 2025 | 6 |
        """
        if not (1 <= month <= 12):
            raise ValueError(f"Invalid month: {month}")
        try:
            end_day = calendar.monthlen(year, month)
            start_date = datetime(year, month, 1)
            end_date = datetime(year, month, end_day)
            return (
                start_date.strftime("%Y-%m-%d 00:00:00.0000"),
                end_date.strftime("%Y-%m-%d 23:59:59.9999")
            )
        except ValueError as e:
            raise ValueError(f"Invalid year or month: {e}")

    def get_days_in_month(self, year: int, month: int) -> int:
        """Get the number of days in a specified month.

        Args:
            year: Year of the month.
            month: Month number (1-12).

        Returns:
            int: Number of days in the month.

        Raises:
            ValueError: If month is invalid.

        Examples:
            | ${days}= | Get Days In Month | 2025 | 6 |
        """
        if not (1 <= month <= 12):
            raise ValueError(f"Invalid month: {month}")
        return calendar.monthlen(year, month)

    def is_divisible_with_no_remainder(self, first_number: Union[int, str], second_number: Union[int, str]) -> bool:
        """Check if first_number is divisible by second_number with no remainder.

        Args:
            first_number: The dividend.
            second_number: The divisor.

        Returns:
            bool: True if divisible with no remainder, False otherwise.

        Raises:
            ValueError: If inputs cannot be converted to integers.

        Examples:
            | ${result}= | Is Divisible With No Remainder | 10 | 2 |
        """
        try:
            return int(first_number) % int(second_number) == 0
        except (ValueError, TypeError) as e:
            raise ValueError(f"Invalid numbers: {e}")

    def is_not_number(self, value: Any) -> bool:
        """Check if a value is not a number.

        Args:
            value: Value to check.

        Returns:
            bool: True if value is not a number, False otherwise.

        Examples:
            | ${is_not_num}= | Is Not Number | abc |
        """
        try:
            float(value)
            return False
        except (ValueError, TypeError):
            return True

    def get_random_number_in_range(self, start: Union[int, str], end: Union[int, str]) -> int:
        """Get a random integer within the specified range (inclusive).

        Args:
            start: Start of the range.
            end: End of the range.

        Returns:
            int: Random integer.

        Raises:
            ValueError: If start or end is invalid or start > end.

        Examples:
            | ${number}= | Get Random Number | 1 | 10 |
        """
        try:
            start_num, end_num = int(start), int(end)
            if start_num > end_num:
                raise ValueError(f"Start ({start_num}) must be less than or equal to end ({end_num})")
            return random.randint(start_num, end_num)
        except (ValueError, TypeError) as e:
            raise ValueError(f"Invalid range: {e}")

    def compare_values(self, value1: Any, relation: str, value2: Any) -> bool:
        """Compare two values using the specified relation.

        Args:
            value1: First value.
            relation: Comparison operator ('<', '>', '<=', '>=', '=').
            value2: Second value.

        Returns:
            bool: Result of the comparison.

        Raises:
            ValueError: If relation is invalid or values cannot be compared.

        Examples:
            | ${result}= | Compare Values | 10 | > | 5 |
        """
        valid_relations = {'<', '>', '<=', '>=', '='}
        if relation not in valid_relations:
            raise ValueError(f"Invalid relation '{relation}'. Supported: {valid_relations}")
        try:
            v1, v2 = float(self.parse_string(value1)), float(self.parse_string(value2))
            comparisons = {
                '<': v1 < v2,
                '>': v1 > v2,
                '<=': v1 <= v2,
                '>=': v1 >= v2,
                '=': v1 == v2
            }
            return comparisons[relation]
        except (ValueError, TypeError) as e:
            raise ValueError(f"Cannot compare '{value1}' and '{value2}': {e}")

    def parse_string(self, value: str) -> Any:
        """Convert a string to its appropriate type (e.g., int, float, list).

        Args:
            value: String to parse.

        Returns:
            Any: Parsed value or original string if parsing fails.

        Examples:
            | ${result}= | Parse String | 123 |
        """
        try:
            return ast.literal_eval(str(value))
        except (ValueError, SyntaxError):
            return value

    def is_list_sorted_ascending(self, values: List[Any]) -> bool:
        """Check if a list is sorted in ascending order.

        Args:
            values: List to check.

        Returns:
            bool: True if sorted ascending, False otherwise.

        Examples:
            | ${is_sorted}= | Is List Sorted Ascending | ${list} |
        """
        return all(self.parse_string(a) <= self.parse_string(b) for a, b in zip(values, values[1:]))

    def is_list_sorted_descending(self, values: List[Any]) -> bool:
        """Check if a list is sorted in descending order.

        Args:
            values: List to check.

        Returns:
            bool: True if sorted descending, False otherwise.

        Examples:
            | ${is_sorted}= | Is List Sorted Descending | ${list} |
        """
        return all(self.parse_string(a) >= self.parse_string(b) for a, b in zip(values, values[1:]))

    def is_list_sorted_equal_or_ascending(self, input_list: List[Any]) -> bool:
        """Check if a list is sorted in ascending order or equal.

        Args:
            input_list: List to check.

        Returns:
            bool: True if sorted ascending or equal, False otherwise.

        Examples:
            | ${is_sorted}= | Is List Sorted Equal Or Ascending | ${list} |
        """
        return all(input_list[i] <= input_list[i + 1] for i in range(len(input_list) - 1))

    def is_list_sorted_equal_or_descending(self, input_list: List[Any]) -> bool:
        """Check if a list is sorted in descending order or equal.

        Args:
            input_list: List to check.

        Returns:
            bool: True if sorted descending or equal, False otherwise.

        Examples:
            | ${is_sorted}= | Is List Sorted Equal Or Descending | ${list} |
        """
        return all(input_list[i] >= input_list[i + 1] for i in range(len(input_list) - 1))

    def is_empty_structure(self, structure: Any) -> bool:
        """Check if a structure (list, dict, set, string, tuple) is empty.

        Args:
            structure: Structure to check.

        Returns:
            bool: True if empty, False otherwise.

        Examples:
            | ${is_empty}= | Is Empty Structure | ${list} |
        """
        return not bool(structure)

    def is_minimum_value(self, values: List[Any], value: Any) -> bool:
        """Check if a value is less than or equal to the minimum in a list.

        Args:
            values: List of values.
            value: Value to compare.

        Returns:
            bool: True if value is less than or equal to minimum, False otherwise.

        Raises:
            ValueError: If list is empty or values cannot be compared.

        Examples:
            | ${is_min}= | Is Minimum Value | ${list} | 5 |
        """
        if not values:
            raise ValueError("List cannot be empty")
        try:
            min_value = min(values, key=float)
            return self.parse_string(value) <= self.parse_string(min_value)
        except (ValueError, TypeError) as e:
            raise ValueError(f"Cannot find minimum in list: {e}")

    def is_maximum_value(self, values: List[Any], value: Any) -> bool:
        """Check if a value is greater than or equal to the maximum in a list.

        Args:
            values: List of values.
            value: Value to compare.

        Returns:
            bool: True if value is greater than or equal to maximum, False otherwise.

        Raises:
            ValueError: If list is empty or values cannot be compared.

        Examples:
            | ${is_max}= | Is Maximum Value | ${list} | 10 |
        """
        if not values:
            raise ValueError("List cannot be empty")
        try:
            max_value = max(values, key=float)
            return self.parse_string(value) >= self.parse_string(max_value)
        except (ValueError, TypeError) as e:
            raise ValueError(f"Cannot find maximum in list: {e}")

    def is_sublist(self, sublist: List[Any], main_list: List[Any]) -> bool:
        """Check if one list is a sublist of another.

        Args:
            sublist: Potential sublist.
            main_list: Main list.

        Returns:
            bool: True if sublist is contained in main_list, False otherwise.

        Examples:
            | ${is_sublist}= | Is Sublist | ${sublist} | ${main_list} |
        """
        return set(sublist).issubset(set(main_list))

    def are_lists_equal(self, list1: List[Any], list2: List[Any]) -> bool:
        """Check if two lists contain the same elements (ignoring order).

        Args:
            list1: First list.
            list2: Second list.

        Returns:
            bool: True if lists are equal, False otherwise.

        Examples:
            | ${are_equal}= | Are Lists Equal | ${list1} | ${list2} |
        """
        return set(list1) == set(list2)

    def concatenate_lists(self, list1: List[Any], list2: List[Any]) -> List[Any]:
        """Concatenate two lists.

        Args:
            list1: First list.
            list2: Second list.

        Returns:
            List: Concatenated list.

        Examples:
            | ${result}= | Concatenate Lists | ${list1} | ${list2} |
        """
        return list1 + list2

    def difference_lists(self, list1: List[Any], list2: List[Any]) -> List[Any]:
        """Get elements in list1 that are not in list2.

        Args:
            list1: First list.
            list2: Second list.

        Returns:
            List: Difference of lists.

        Examples:
            | ${diff}= | Difference Lists | ${list1} | ${list2} |
        """
        return list(set(list1) - set(list2))

    def sum_lists(self, list1: List[Any], list2: List[Any]) -> List[Any]:
        """Sum corresponding elements of two lists.

        Args:
            list1: First list.
            list2: Second list.

        Returns:
            List: List of sums.

        Raises:
            ValueError: If lists have different lengths or elements cannot be summed.

        Examples:
            | ${sums}= | Sum Lists | ${list1} | ${list2} |
        """
        if len(list1) != len(list2):
            raise ValueError(f"Lists must have equal length: {len(list1)} != {len(list2)}")
        try:
            return [self.parse_string(a) + self.parse_string(b) for a, b in zip(list1, list2)]
        except (ValueError, TypeError) as e:
            raise ValueError(f"Cannot sum lists: {e}")

    def subtract_lists(self, list1: List[Any], list2: List[Any]) -> List[Any]:
        """Subtract corresponding elements of two lists.

        Args:
            list1: First list.
            list2: Second list.

        Returns:
            List: List of differences.

        Raises:
            ValueError: If lists have different lengths or elements cannot be subtracted.

        Examples:
            | ${diffs}= | Subtract Lists | ${list1} | ${list2} |
        """
        if len(list1) != len(list2):
            raise ValueError(f"Lists must have equal length: {len(list1)} != {len(list2)}")
        try:
            return [self.parse_string(a) - self.parse_string(b) for a, b in zip(list1, list2)]
        except (ValueError, TypeError) as e:
            raise ValueError(f"Cannot subtract lists: {e}")

    def get_common_values(self, list1: List[Any], list2: List[Any]) -> List[Any]:
        """Get common elements between two lists.

        Args:
            list1: First list.
            list2: Second list.

        Returns:
            List: List of common elements.

        Examples:
            | ${common}= | Get Common Values | ${list1} | ${list2} |
        """
        return list(set(list1) & set(list2))

    def remove_empty_strings(self, input_list: List[str]) -> List[str]:
        """Remove empty strings from a list.

        Args:
            input_list: List containing strings.

        Returns:
            List: List with empty strings removed.

        Examples:
            | ${cleaned}= | Remove Empty Strings | ${list} |
        """
        return [item for item in input_list if item]

    def are_all_values_equal(self, input_list: List[Any], value: Any) -> bool:
        """Check if all values in a list equal a given value.

        Args:
            input_list: List to check.
            value: Value to compare against.

        Returns:
            bool: True if all values match, False otherwise.

        Examples:
            | ${all_equal}= | Are All Values Equal | ${list} | 5 |
        """
        return all(self.parse_string(value) == self.parse_string(item) for item in input_list)

    def do_all_values_contain(self, input_list: List[Any], substring: Any) -> bool:
        """Check if all values in a list contain a substring.

        Args:
            input_list: List to check.
            substring: Substring to search for.

        Returns:
            bool: True if all values contain substring, False otherwise.

        Examples:
            | ${all_contain}= | Do All Values Contain | ${list} | test |
        """
        return all(str(self.parse_string(substring)) in str(self.parse_string(item)) for item in input_list)

    def get_minimum_value(self, values: List[Any]) -> Any:
        """Get the minimum value from a list.

        Args:
            values: List of values.

        Returns:
            Any: Minimum value.

        Raises:
            ValueError: If list is empty.

        Examples:
            | ${min}= | Get Minimum Value | ${list} |
        """
        if not values:
            raise ValueError("List cannot be empty")
        return min(values, key=float)

    def get_maximum_value(self, values: List[Any]) -> Any:
        """Get the maximum value from a list.

        Args:
            values: List of values.

        Returns:
            Any: Maximum value.

        Raises:
            ValueError: If list is empty.

        Examples:
            | ${max}= | Get Maximum Value | ${list} |
        """
        if not values:
            raise ValueError("List cannot be empty")
        return max(values, key=float)

    def flatten_list(self, nested_list: List[Any]) -> List[Any]:
        """Flatten a nested list.

        Args:
            nested_list: Nested list to flatten.

        Returns:
            List: Flattened list.

        Examples:
            | ${flat}= | Flatten List | ${nested_list} |
        """
        return [item for sublist in nested_list for item in (sublist if isinstance(sublist, list) else [sublist])]

    def count_items_in_list(self, input_list: List[Any]) -> Dict[Any, int]:
        """Count occurrences of items in a list.

        Args:
            input_list: List to count.

        Returns:
            Dict: Dictionary with items and their counts.

        Examples:
            | ${counts}= | Count Items In List | ${list} |
        """
        return dict(Counter(input_list))

    def get_items_starting_with(self, input_list: List[str], prefix: str) -> List[str]:
        """Get items from a list that start with a prefix.

        Args:
            input_list: List of strings.
            prefix: Prefix to match.

        Returns:
            List[str]: List of matching items.

        Examples:
            | ${items}= | Get Items Starting With | ${list} | temp |
        """
        return [item for item in input_list if item.startswith(prefix)]

    def does_string_contain(self, string1: str, string2: str) -> bool:
        """Check if string1 contains string2.

        Args:
            string1: String to search in.
            string2: Substring to search for.

        Returns:
            bool: True if string2 is in string1, False otherwise.

        Examples:
            | ${contains}= | Does String Contain | hello world | world |
        """
        return str(string2) in str(string1)

    def clean_string(self, string: str) -> str:
        """Remove single quotes from a string.

        Args:
            string: String to clean.

        Returns:
            str: Cleaned string.

        Examples:
            | ${cleaned}= | Clean String | 'hello' |
        """
        return str(string).replace("'", "")

    def format_search_string(self, value: Union[str, List[str]]) -> str:
        """Format a value or list for search (join lists with commas).

        Args:
            value: String or list to format.

        Returns:
            str: Formatted string.

        Examples:
            | ${formatted}= | Format Search String | ${list} |
        """
        return ','.join(value) if isinstance(value, list) else str(value)

    def encode_to_base64(self, string: str) -> str:
        """Encode a string to base64.

        Args:
            string: String to encode.

        Returns:
            str: Base64 encoded string.

        Raises:
            TypeError: If input is not a string.

        Examples:
            | ${encoded}= | Encode To Base64 | hello |
        """
        if not isinstance(string, str):
            raise TypeError(f"Input must be a string, got {type(string)}")
        return base64.b64encode(string.encode()).decode()

    def decode_from_base64(self, string: str) -> str:
        """Decode a base64 string to plain text.

        Args:
            string: Base64 encoded string.

        Returns:
            str: Decoded string.

        Raises:
            ValueError: If decoding fails.

        Examples:
            | ${decoded}= | Decode From Base64 | aGVsbG8= |
        """
        try:
            return base64.b64decode(string).decode()
        except (ValueError, TypeError) as e:
            raise ValueError(f"Invalid base64 string: {e}")

    def is_valid_json(self, string: str) -> bool:
        """Check if a string is valid JSON.

        Args:
            string: String to check.

        Returns:
            bool: True if valid JSON, False otherwise.

        Examples:
            | ${is_json}= | Is Valid JSON | {"key": "value"} |
        """
        try:
            json.loads(str(string))
            return True
        except ValueError:
            return False

    def format_json_with_values(self, keys: Sequence[str], template: str, key_name: str = "code") -> str:
        """Format a JSON template with a list of values.

        Args:
            keys: List of values to insert.
            template: JSON template string (e.g., '{"id": "837"}').
            key_name: Key to replace in the template.

        Returns:
            str: JSON string with formatted values.

        Raises:
            ValueError: If template is invalid JSON.

        Examples:
            | ${json}= | Format JSON With Values | ${keys} | {"id": "837"} | id |
        """
        try:
            tmpl = json.loads(template)
            return json.dumps([{**tmpl, key_name: k} for k in keys])
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON template: {e}")

    def format_dict_with_values(self, keys: Sequence[str], template: str, key_name: str = "code") -> List[Dict]:
        """Format a dictionary template with a list of values.

        Args:
            keys: List of values to insert.
            template: JSON template string (e.g., '{"id": "837"}').
            key_name: Key to replace in the template.

        Returns:
            List[Dict]: List of formatted dictionaries.

        Raises:
            ValueError: If template is invalid JSON.

        Examples:
            | ${dicts}= | Format Dict With Values | ${keys} | {"id": "837"} | id |
        """
        try:
            tmpl = json.loads(template)
            return [{**tmpl, key_name: k} for k in keys]
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON template: {e}")

    def overwrite_json_template(
        self,
        template: str,
        key1: str,
        values1: List[Any],
        key2: str,
        values2: List[Any],
        key3: str,
        values3: List[Any]
    ) -> str:
        """Overwrite a JSON template with multiple lists of values.

        Args:
            template: JSON template string.
            key1: First key to replace.
            values1: First list of values.
            key2: Second key to replace.
            values2: Second list of values.
            key3: Third key to replace.
            values3: Third list of values.

        Returns:
            str: JSON string with formatted values.

        Raises:
            ValueError: If template is invalid JSON or lists have different lengths.

        Examples:
            | ${json}= | Overwrite JSON Template | {"data": {}} | k1 | ${list1} | k2 | ${list2} | k3 | ${list3} |
        """
        try:
            tmpl = json.loads(template)
            if not (len(values1) == len(values2) == len(values3)):
                raise ValueError("All input lists must have the same length")
            return json.dumps([{**tmpl, key1: v1, key2: values2[idx], key3: values3[idx]} for idx, v1 in enumerate(values1)])
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON template: {e}")

    def format_json_with_loop_numbers(self, code: str, count: int, template: str, key_name: str = "orderCode") -> List[Dict]:
        """Format a JSON template with sequential numbers appended to a code.

        Args:
            code: Base code string.
            count: Number of items to generate.
            template: JSON template string.
            key_name: Key to replace in the template.

        Returns:
            List[Dict]: List of formatted dictionaries.

        Raises:
            ValueError: If template is invalid JSON or count is negative.

        Examples:
            | ${json}= | Format JSON With Loop Numbers | ABC | 3 | {"id": "837"} | orderCode |
        """
        if count < 0:
            raise ValueError("Count must be non-negative")
        try:
            tmpl = json.loads(template)
            return [{**tmpl, key_name: f"{code}{k}"} for k in range(count)]
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON template: {e}")

    def read_json_file(self, filename: str) -> Dict:
        """Read a JSON file and return its contents.

        Args:
            filename: Path to the JSON file.

        Returns:
            Dict: Parsed JSON data.

        Raises:
            FileNotFoundError: If file does not exist.
            ValueError: If file is not valid JSON.

        Examples:
            | ${data}= | Read JSON File | data.json |
        """
        try:
            with open(filename, encoding="utf8") as file:
                return json.load(file)
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {filename}")
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in file {filename}: {e}")

    def create_directory_if_not_exists(self, directory: str) -> None:
        """Create a directory if it does not exist.

        Args:
            directory: Path to the directory.

        Examples:
            | Create Directory If Not Exists | /path/to/dir |
        """
        os.makedirs(directory, exist_ok=True)

    def get_latest_screenshot_path(self, directory: str, is_web_screenshot: bool) -> str:
        """Get the path of the latest screenshot file in the directory.

        Args:
            directory: Directory to search for screenshots.
            is_web_screenshot: True for Selenium screenshots, False for Appium.

        Returns:
            str: Path to the latest screenshot file.

        Raises:
            ValueError: If no screenshot is found or directory is invalid.

        Examples:
            | ${path}= | Get Latest Screenshot Path | /screenshots | True |
        """
        if not os.path.isdir(directory):
            raise ValueError(f"Invalid directory: {directory}")
        prefix = "selenium" if is_web_screenshot else "appium"
        files = glob.glob(os.path.join(directory, f"{prefix}*.png"))
        if not files:
            raise ValueError(f"No {prefix} screenshots found in {directory}")
        return max(files, key=os.path.getmtime)

    def get_file_path(self, filename: str) -> str:
        """Get the full path to a file in the current directory.

        Args:
            filename: Name of the file.

        Returns:
            str: Full path to the file.

        Examples:
            | ${path}= | Get File Path | example.txt |
        """
        return os.path.join(os.getcwd(), filename)

    def list_files_in_directory(self, directory: str) -> List[str]:
        """List all files in a directory.

        Args:
            directory: Path to the directory.

        Returns:
            List[str]: List of file names.

        Raises:
            FileNotFoundError: If directory does not exist.

        Examples:
            | ${files}= | List Files In Directory | /path/to/dir |
        """
        if not os.path.isdir(directory):
            raise FileNotFoundError(f"Directory not found: {directory}")
        return os.listdir(directory)
    
    def count_files_in_zip(self, zip_file):
        """
        Đếm số lượng tệp trong tệp zip mà không cần giải nén.
        
        Args:
            zip_file (str): Đường dẫn đến tệp zip
        
        Returns:
            int: Số lượng tệp trong tệp zip
        """
        try:
            # Check if zip file exists
            if not os.path.exists(zip_file):
                logger.error(f"Zip file does not exist: {zip_file}")
                raise Exception(f"Zip file does not exist: {zip_file}")
            
            with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                if zip_ref.testzip() is not None:
                    logger.error(f"Corrupted zip file: {zip_file}")
                    raise Exception(f"Corrupted zip file: {zip_file}")
                
                file_count = len([f for f in zip_ref.namelist() if not f.endswith('/')])
            
            logger.info(f"Number of files in {zip_file}: {file_count}")
            return file_count

        except zipfile.BadZipFile:
            logger.error(f"Invalid zip file: {zip_file}")
            raise Exception(f"Invalid zip file: {zip_file}")
        except Exception as e:
            logger.error(f"Error: {str(e)}")
            raise
    def create_excel_sheet(self, file_path: str, sheet_name: str) -> None:
        """Create a new sheet in an Excel file.

        Args:
            file_path: Path to the Excel file.
            sheet_name: Name of the new sheet.

        Raises:
            FileNotFoundError: If file does not exist.
            ValueError: If sheet_name is invalid.

        Examples:
            | Create Excel Sheet | /path/to/file.xlsx | Sheet1 |
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        if not sheet_name:
            raise ValueError("Sheet name cannot be empty")
        try:
            workbook = load_workbook(file_path, data_only=True)
            workbook.create_sheet(title=sheet_name)
            workbook.save(file_path)
        except Exception as e:
            raise ValueError(f"Failed to create sheet: {e}")

    def switch_browser_tab(self, tab_index: Union[int, str]) -> None:
        """Switch to a browser tab by index.

        Args:
            tab_index: Index of the tab to switch to (0-based).

        Raises:
            ValueError: If tab_index is invalid or out of range.

        Examples:
            | Switch Browser Tab | 1 |
        """
        try:
            tab_index = int(tab_index)
            driver = BuiltIn().get_library_instance("SeleniumLibrary")._current_browser()
            if tab_index < 0 or tab_index >= len(driver.window_handles):
                raise ValueError(f"Invalid tab index: {tab_index}")
            driver.switch_to.window(driver.window_handles[tab_index])
        except (ValueError, TypeError) as e:
            raise ValueError(f"Failed to switch tab: {e}")

    def compare_images(self, first_image_path: str, second_image_path: str) -> float:
        """Compare two images and return the difference percentage.

        Args:
            first_image_path: Path to the first image.
            second_image_path: Path to the second image.

        Returns:
            float: Difference percentage between the images.

        Raises:
            FileNotFoundError: If images do not exist.
            ValueError: If images have different modes or sizes.

        Examples:
            | ${diff}= | Compare Images | /img1.png | /img2.png |
        """
        for path in (first_image_path, second_image_path):
            if not os.path.isfile(path):
                raise FileNotFoundError(f"Image not found: {path}")
        try:
            img1 = Image.open(first_image_path.replace('\\', '/'))
            img2 = Image.open(second_image_path.replace('\\', '/'))
            if img1.mode != img2.mode:
                raise ValueError("Images have different modes")
            if img1.size != img2.size:
                raise ValueError("Images have different sizes")
            pairs = zip_longest(img1.getdata(), img2.getdata())
            diff = sum(abs(c1 - c2) for p1, p2 in pairs for c1, c2 in zip(p1, p2)) if img1.getbands() != 1 else \
                   sum(abs(p1 - p2) for p1, p2 in pairs)
            n_components = img1.size[0] * img1.size[1] * (3 if img1.getbands() != 1 else 1)
            return (diff / 255.0 * 100) / n_components
        except Exception as e:
            raise ValueError(f"Failed to compare images: {e}")

    def rename_dictionary_keys(self, dictionary: Dict, new_keys: List[str], old_keys: List[str]) -> Dict:
        """Rename keys in a dictionary.

        Args:
            dictionary: Dictionary to modify.
            new_keys: List of new keys.
            old_keys: List of old keys.

        Returns:
            Dict: Updated dictionary.

        Raises:
            ValueError: If lengths of new_keys and old_keys do not match.

        Examples:
            | ${new_dict}= | Rename Dictionary Keys | ${dict} | ${new_keys} | ${old_keys} |
        """
        if len(new_keys) != len(old_keys):
            raise ValueError("New keys and old keys must have the same length")
        for new_key, old_key in zip(new_keys, old_keys):
            if old_key in dictionary:
                dictionary[new_key] = dictionary.pop(old_key)
        return dictionary

    def create_dictionary_without_empty_values(self, *items: Any) -> Dict:
        """Create a dictionary and remove keys with empty values.

        Args:
            *items: Key-value pairs.

        Returns:
            Dict: Dictionary without empty values.

        Examples:
            | ${dict}= | Create Dictionary Without Empty Values | key1=value1 | key2= |
        """
        from robot.libraries.BuiltIn import _Converter, _BuiltInBase
        from robot.variables import DictVariableResolver

        converter = _Converter()
        var = _BuiltInBase()
        separate, combined = converter._split_dict_items(items)
        result = DotDict(converter._format_separate_dict_items(separate))
        combined = DictVariableResolver(combined).resolve(var._variables)
        result.update(combined)
        return {k: v for k, v in result.items() if v}

    def generate_otp(self, key_token: str) -> str:
        """Generate a TOTP based on the provided key.

        Args:
            key_token: TOTP secret key.

        Returns:
            str: Generated OTP.

        Raises:
            ValueError: If key_token is invalid.

        Examples:
            | ${otp}= | Generate OTP | V2RHBVUWSHCQZPZIC3ZC2GORYRBIQTNP |
        """
        try:
            totp = pyotp.TOTP(key_token)
            return totp.now()
        except Exception as e:
            raise ValueError(f"Invalid TOTP key: {e}")
    
    def Gen_uuid(self):
        uid =  uuid.uuid4()
        return uid
    
    def load_test_data(self, test_case_name, file_name):
        """Load test data from a YAML file based on the test case name and environment."""

        env = BuiltIn().get_variable_value('${env}')
        print(f"Loading test data for {test_case_name} in environment {env} from {file_name}")
        with open(f"{file_name}", "r", encoding="utf-8") as file:
            data = yaml.safe_load(file)
        print(f"Data loaded: {data}")
        return data["test_cases"][test_case_name][env]    
        
    
    def fix_excel_real(self, file_path):
        """Fix Excel file trong các trường hợp update dữ liệu bằng code làm Excel không tự recalculation hoặc gây lỗi reference giữa các sheet"""
        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False

        try:
            wb = app.books.open(file_path, update_links=False)

            wb.api.Application.CalculateFullRebuild()

            wb.save()
            wb.close()
        finally:
            app.quit()